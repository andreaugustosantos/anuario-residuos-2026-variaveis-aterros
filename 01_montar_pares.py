"""
01 - Montagem dos pares origem-destino.

Lê o SINISA e o shapefile do IBGE e produz, para cada municipio do Anuario, o par de coordenadas entre o qual a rota será calculada.

Regra de seleção do destino: entre as rotas cuja unidade está identificada no cadastro e possui coordenada disponível, adota-se a de maior massa anual
(GTR1008). A unidade dessa rota (GTR1013) define o ponto de chegada.

Exclusoes aplicadas:
  - município sem rota de coleta declarada;
  - rota sem identificação da unidade de destino;
  - unidade fora do cadastro de infraestrutura;
  - coordenada da unidade coincidente com o ponto do município que a abriga
    (preenchimento padrao do formulário, não posição real da instalação).

Saída: saida/01_pares.csv"""

import csv
import math
import sys
from collections import defaultdict

import openpyxl
import shapefile

import config


# APOIO


def haversine(a, b):
    """Distância em linha reta, em quilômetros, entre dois pares (lat, lon)."""
    raio = 6371.0088
    lat1, lon1 = map(math.radians, a)
    lat2, lon2 = map(math.radians, b)
    return 2 * raio * math.asin(math.sqrt(
        math.sin((lat2 - lat1) / 2) ** 2
        + math.cos(lat1) * math.cos(lat2) * math.sin((lon2 - lon1) / 2) ** 2))


def numero(valor):
    try:
        return float(str(valor).replace(",", "."))
    except (TypeError, ValueError):
        return None


def coordenada(valor):
    """Extrai o numero de textos como 'Lat -9.9350712'."""
    if valor is None:
        return None
    texto = str(valor).replace(",", ".")
    numeros = "".join(c for c in texto if c in "-0123456789.")
    try:
        return float(numeros)
    except ValueError:
        return None


def exige(caminho):
    if not caminho.exists():
        sys.exit(
            "Nao encontrei:\n  %s\n\n"
            "Confira os caminhos em config.py e o README para saber onde baixar."
            % caminho
        )



# LEITURA DAS FONTES


def ler_municipios():
    """Os 2.000 municípios do Anuario, do CSV que acompanha o repositório."""
    exige(config.MUNICIPIOS)
    with open(config.MUNICIPIOS, encoding="utf-8") as f:
        return [linha for linha in csv.DictReader(f)]


def ler_coordenadas_referencia():
    """Tabela de coordenadas municipais usada apenas no controle de qualidade."""
    exige(config.COORDENADAS_REFERENCIA)
    with open(config.COORDENADAS_REFERENCIA, encoding="utf-8") as f:
        return {
            linha["codigo_ibge"].strip():
                (float(linha["latitude"]), float(linha["longitude"]))
            for linha in csv.DictReader(f)
        }


def ler_sedes():
    """Coordenada da sede de cada município, do shapefile do IBGE."""
    exige(config.SHAPEFILE.with_suffix(".shp"))
    subcategorias = {"Sede Municipal", "Capital Estadual", "Capital Federal"}
    sedes = {}
    leitor = shapefile.Reader(str(config.SHAPEFILE), encoding="utf-8")
    for registro in leitor.iterShapeRecords():
        r = registro.record
        categoria = r["CT_LOCALID"]
        if categoria == "Cidade":
            if r["SCT_LOCALI"] not in subcategorias:
                continue
        elif categoria != "Distrito Estadual de Fernando de Noronha":
            continue
        codigo = str(r["CD_MUN"]).strip()
        lon, lat = registro.shape.points[0]
        # capitais aparecem duas vezes; a Sede Municipal tem precedencia
        if codigo in sedes and r["SCT_LOCALI"] != "Sede Municipal":
            continue
        sedes[codigo] = (lat, lon)
    return sedes


def ler_unidades():
    """Cadastro das unidades de destinação, com coordenada declarada."""
    formularios = [
        (config.DISPOSICAO_FINAL, "Infra_Disposição_Final_caracter", 16, 18, 19, 20),
        (config.PROCESSAMENTO, "Infra_Proces_Trat_caracteristic", 16, 18, 19, 20),
        (config.TRANSBORDO, "Infra_Transbordo_caracteristica", 16, 17, 18, 19),
    ]
    unidades = {}
    for caminho, aba, col_codigo, col_nome, col_lat, col_lon in formularios:
        exige(caminho)
        planilha = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
        for linha in planilha[aba].iter_rows(min_row=config.PRIMEIRA_LINHA,
                                             values_only=True):
            codigo = linha[col_codigo]
            if not codigo:
                continue
            unidades[str(codigo).strip()] = {
                "municipio": str(linha[1]).strip() if linha[1] else None,
                "nome": linha[col_nome],
                "lat": coordenada(linha[col_lat]),
                "lon": coordenada(linha[col_lon]),
            }
        planilha.close()
    return unidades


def ler_rotas():
    """
    Rotas de coleta declaradas por municipio, do formulario Manejo.
    """
    exige(config.MANEJO)
    planilha = openpyxl.load_workbook(config.MANEJO, read_only=True, data_only=True)
    rotas = defaultdict(list)
    for linha in planilha["Manejo_Coleta_e_Destinação"].iter_rows(
            min_row=config.PRIMEIRA_LINHA, values_only=True):
        codigo_municipio = linha[1]
        unidade = linha[30]
        if not codigo_municipio or not unidade:
            continue
        rotas[str(codigo_municipio).strip()].append({
            "massa": numero(linha[24]),
            "codigo_unidade": str(unidade).split(" - ")[0].strip(),
            "tipo": linha[28],
        })
    planilha.close()
    return rotas


# MONTAGEM

# Colunas que este script acrescenta as do arquivo de municipios.
COLUNAS_PARES = [
    "origem_lat", "origem_lon", "destino_lat", "destino_lon",
    "codigo_unidade", "nome_unidade", "tipo_unidade", "municipio_da_unidade",
    "fluxo", "massa_rota_t_ano", "dist_linha_reta_km",
]
COLUNAS_EXCLUIDOS = [
    "motivo", "codigo_unidade", "nome_unidade", "tipo_unidade",
    "municipio_da_unidade",
]

def rota_dominante(rotas_do_municipio, unidades):
    """A rota de maior massa entre as que apontam unidade com coordenada."""
    candidatas = [
        r for r in rotas_do_municipio
        if r["codigo_unidade"] in unidades
        and unidades[r["codigo_unidade"]]["lat"] is not None
        and unidades[r["codigo_unidade"]]["lon"] is not None
    ]
    if not candidatas:
        return None
    com_massa = [r for r in candidatas if r["massa"] is not None]
    return max(com_massa, key=lambda r: r["massa"]) if com_massa else candidatas[0]


def motivo_da_exclusao(rotas_do_municipio):
    if not rotas_do_municipio:
        return "Sem rota de coleta declarada no SINISA"
    com_massa = [r for r in rotas_do_municipio if r["massa"] is not None]
    escolhida = (max(com_massa, key=lambda r: r["massa"])
                 if com_massa else rotas_do_municipio[0])
    codigo = escolhida["codigo_unidade"]
    if codigo == "Null":
        return "Rota declarada sem identificacao da unidade de destino"
    if "Outro" in codigo:
        return "Unidade de destino fora do cadastro SINISA"
    return "Unidade de destino sem coordenada declarada"


def e_preenchimento_padrao(unidade, referencias):
    """Testa se a coordenada declarada reproduz o ponto do municipio que abriga a unidade, em qualquer uma das tabelas de referencia recebidas."""
    ponto = (unidade["lat"], unidade["lon"])
    for referencia in referencias:
        if referencia is None:
            continue
        mesma_quarta_casa = (round(ponto[0], 4) == round(referencia[0], 4)
                             and round(ponto[1], 4) == round(referencia[1], 4))
        if mesma_quarta_casa or haversine(ponto, referencia) < config.LIMITE_COINCIDENCIA_KM:
            return True
    return False


def main():
    print("Lendo as bases...")
    municipios = ler_municipios()
    sedes = ler_sedes()
    referencias = ler_coordenadas_referencia()
    unidades = ler_unidades()
    rotas = ler_rotas()
    print("  %d municipios no Anuario" % len(municipios))
    print("  %d sedes municipais no IBGE" % len(sedes))
    print("  %d unidades no cadastro do SINISA" % len(unidades))
    print("  %d municipios com rota declarada" % len(rotas))

    pares, excluidos = [], []
    for m in municipios:
        codigo = m["cod_ibge"]
        escolhida = rota_dominante(rotas.get(codigo, []), unidades)

        if escolhida is None:
            excluidos.append({**m, "motivo": motivo_da_exclusao(rotas.get(codigo, [])),
                              "codigo_unidade": "", "nome_unidade": "",
                              "tipo_unidade": "", "municipio_da_unidade": ""})
            continue

        unidade = unidades[escolhida["codigo_unidade"]]
        anfitriao = unidade["municipio"]

        if e_preenchimento_padrao(unidade, (sedes.get(anfitriao),
                                            referencias.get(anfitriao))):
            excluidos.append({
                **m,
                "motivo": "Coordenada da unidade coincide com o ponto do municipio",
                "codigo_unidade": escolhida["codigo_unidade"],
                "nome_unidade": unidade["nome"],
                "tipo_unidade": escolhida["tipo"],
                "municipio_da_unidade": anfitriao,
            })
            continue

        origem = sedes.get(codigo)
        if origem is None:
            excluidos.append({**m, "motivo": "Sede municipal ausente no shapefile",
                              "codigo_unidade": escolhida["codigo_unidade"],
                              "nome_unidade": unidade["nome"],
                              "tipo_unidade": escolhida["tipo"],
                              "municipio_da_unidade": anfitriao})
            continue

        pares.append({
            **m,
            "origem_lat": round(origem[0], 6),
            "origem_lon": round(origem[1], 6),
            "destino_lat": unidade["lat"],
            "destino_lon": unidade["lon"],
            "codigo_unidade": escolhida["codigo_unidade"],
            "nome_unidade": unidade["nome"],
            "tipo_unidade": escolhida["tipo"],
            "municipio_da_unidade": anfitriao,
            "fluxo": "Intramunicipal" if anfitriao == codigo else "Intermunicipal",
            "massa_rota_t_ano": escolhida["massa"],
            "dist_linha_reta_km": round(
                haversine(origem, (unidade["lat"], unidade["lon"])), 1),
        })

    config.SAIDA.mkdir(exist_ok=True)
    base = list(municipios[0].keys())
    saidas = (("01_pares.csv", pares, base + COLUNAS_PARES),
              ("01_excluidos.csv", excluidos, base + COLUNAS_EXCLUIDOS))
    for nome, dados, campos in saidas:
        with open(config.SAIDA / nome, "w", newline="", encoding="utf-8") as f:
            escritor = csv.DictWriter(f, fieldnames=campos)
            escritor.writeheader()
            escritor.writerows(dados)

    print()
    print("Pares montados ...... %d" % len(pares))
    print("Excluidos ........... %d" % len(excluidos))
    contagem = {}
    for e in excluidos:
        contagem[e["motivo"]] = contagem.get(e["motivo"], 0) + 1
    for motivo, quantidade in sorted(contagem.items(), key=lambda x: -x[1]):
        print("   %4d  %s" % (quantidade, motivo))
    print()
    print("Gravado em %s" % config.SAIDA)

    


if __name__ == "__main__":
    main()
